"""
app/services/geographie.py
==========================
Module Geographie — UC-05, EF-01 a EF-06.

Charge et valide le referentiel enrichi a 5 niveaux depuis
`Loader_Base_FinZuu_v1_1.xlsx`. Ce referentiel est PROPRE AU LOADER : il n'est
jamais pousse dans config-service, qui ne connait que Country.cities[] en texte
libre. Il sert uniquement a distribuer geographiquement l'arbre operationnel.

EF-02 impose la validation de l'integrite referentielle : chaque Region a un
Country parent, chaque City une Region, chaque District une City. UC-05, cas
alternatif, precise le traitement : « le Loader journalise l'incoherence et
exclut les entites enfants dependantes » — on n'interrompt donc pas, on ampute
proprement la branche orpheline.

EF-04 : ajouter un pays ne demande aucune modification de code, uniquement des
lignes dans le fichier source.
"""

from __future__ import annotations

import random
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True, slots=True)
class Pays:
    """Racine du referentiel — le niveau que nous ne chargions pas.

    Le domaine s'enracine sur `Country`, jamais sur Client : c'est la
    hierarchie du CDC §6. Un pays porte sa devise, son indicatif, sa TVA et son
    fuseau ; tout ce qui pend dessous doit s'y accorder.

    **config-service ne connait rien de tout cela** — il expose `Country` avec
    un simple tableau `cities[]` en texte libre. Ce referentiel est donc le
    notre, et il est le seul endroit ou la coherence territoriale existe.

    Champs que nous ignorions et qui servent :

      `dial_code`         l'indicatif. Je le codais en dur dans mes sondes
                          (`+237` pour les 4 pays) — il est ici depuis le debut.
      `devise_iso`        le lien direct pays -> devise, sans passer par
                          `countries_using`.
      `tva_percent`       19,25 % au Cameroun, 18 % en zone UEMOA. La Policy
                          d'un produit Collecte porte un champ `vat` : c'est
                          **ici** qu'il doit venir le chercher, pas d'une
                          constante inventee.
      `regulateur_finance` COBAC en zone CEMAC, BCEAO en UEMOA — c'est le
                          regulateur qui fixe le plafond d'usure d'`EF-35`.
    """

    iso2: str
    nom_fr: str
    nom_en: str
    capitale: str
    dial_code: str
    devise_iso: str
    tva_percent: float
    timezone: str
    region_africa: str
    regulateur_telco: str
    regulateur_finance: str


@dataclass(frozen=True, slots=True)
class Region:
    region_id: str
    country_iso2: str
    name: str
    capitale: str
    population: int | None


@dataclass(frozen=True, slots=True)
class City:
    city_id: str
    country_iso2: str
    region_id: str
    name: str
    poids_economique: float
    #: EF-03 — coordonnees GPS « lorsqu'elles sont disponibles dans le
    #: referentiel ». Seule la feuille Cities les porte : Regions et Districts
    #: n'ont que `population_est`. Elles serviront a deriver les coordonnees des
    #: Address generees (02_class : Address.latitude / Address.longitude).
    latitude: float | None
    longitude: float | None
    population: int | None
    est_capitale_pays: bool
    est_capitale_region: bool


@dataclass(frozen=True, slots=True)
class District:
    """`zone_type` distingue residentiel, commercial, industriel...

    Un Kiosque de quartier residentiel et un Kiosque de quartier commercial
    n'ont pas le meme profil d'activite. Le referentiel le sait ; nous
    l'ignorions.
    """

    district_id: str
    city_id: str
    name: str
    zone_type: str
    population: int | None


@dataclass(frozen=True, slots=True)
class Telco:
    """Operateur telecom, avec sa regle de numerotation reelle.

    `EF-27` exige de valider le MSISDN « contre le regex de l'operateur telco
    du pays ». Le referentiel porte ces 12 regex depuis le depart — nous ne
    les chargions simplement pas.

    `part_marche` est la part de marche **reelle** (MTN CM 46 %, Orange CM
    43 %, Camtel 3 %). Elle sert a repartir les 2000 clients entre operateurs
    comme dans la vie, plutot qu'uniformement : un echantillon ou chaque
    operateur pese un tiers ne ressemble a aucun marche africain.
    """

    telco_id: str
    country_iso2: str
    network_name: str
    short_name: str
    ussd_base_code: str
    regex_msisdn: str
    part_marche: float

    def accepte(self, msisdn: str) -> bool:
        return re.match(self.regex_msisdn, str(msisdn).strip()) is not None

    def composer_msisdn(self, chiffres: str, alea: random.Random) -> str:
        """Compose un MSISDN **conforme au plan de numerotation reel**, en
        utilisant `chiffres` comme matiere.

        **Pourquoi cette fonction existe.** Faker fournit huit chiffres — le
        corps du numero — mais **jamais le prefixe operateur**, qui est
        implicite pour un habitant du pays. Au Cameroun, tout mobile commence
        par `6` ; en Cote d'Ivoire par `01`, `05` ou `07` ; au Burkina par `0`.
        Personne ne le dit, parce que personne ne l'ignore sur place.

        Mesure du 09/08 : sans ce prefixe, **aucun** numero de Faker n'est
        attribuable a un reseau reel. Avec lui, la conformite est totale.

        Le corps vient de Faker, donc **deux executions sur le meme client
        produisent le meme numero** — la tracabilite est preservee (`ENF-15`).
        Seul le choix de la variante, quand le plan en offre plusieurs, est
        tire par `alea`, lui-meme derive du `run_id`.
        """
        corps = re.sub(r"\D", "", str(chiffres))
        variantes = _variantes_du_motif(self.regex_msisdn)
        if not variantes:
            raise ValueError(f"{self.telco_id} : motif de numerotation inexploitable")

        prefixe_pays, segments = alea.choice(variantes)
        numero = prefixe_pays
        curseur = 0
        for segment in segments:
            if isinstance(segment, str):
                numero += segment
                continue
            autorises, taille = segment
            for _ in range(taille):
                source = int(corps[curseur % len(corps)]) if corps else alea.randint(0, 9)
                if autorises is None:
                    numero += str(source)
                else:
                    # Classe restreinte : on ramene le chiffre de Faker dans
                    # l'ensemble autorise plutot que de le remplacer. La
                    # matiere reste la sienne, seule la contrainte est notre.
                    numero += str(autorises[source % len(autorises)])
                curseur += 1

        if not self.accepte(numero):  # pragma: no cover — garde-fou
            raise ValueError(f"{self.telco_id} : numero compose '{numero}' non conforme")
        return numero


#: Un segment de motif : soit des chiffres litteraux, soit un couple
#: (chiffres autorises, repetitions) — `None` autorisant les dix chiffres.
_Segment = str | tuple[tuple[int, ...] | None, int]


def _chiffres_de_la_classe(contenu: str) -> tuple[int, ...]:
    """Developpe le contenu d'une classe `[...]` en chiffres autorises.

    Les 12 motifs du referentiel utilisent **les deux formes** :
    des plages au Cameroun (`[0-4]`, `[5-9]`) et des enumerations au Burkina
    (`[56]`, `[12]`, `[678]`). Les deux doivent etre couvertes.
    """
    autorises: list[int] = []
    position = 0
    while position < len(contenu):
        if position + 2 < len(contenu) and contenu[position + 1] == "-":
            autorises.extend(range(int(contenu[position]), int(contenu[position + 2]) + 1))
            position += 3
        else:
            autorises.append(int(contenu[position]))
            position += 1
    return tuple(dict.fromkeys(autorises))


def _variantes_du_motif(motif: str) -> list[tuple[str, list[_Segment]]]:
    """Decompose un motif `^<indicatif>(<alt>|<alt>)$` en variantes exploitables.

    La grammaire des 12 motifs du referentiel est volontairement etroite :
    chiffres litteraux, classes `[...]` (plages **et** enumerations), et
    `\\d{n}`. On ne cherche pas a couvrir les expressions regulieres en
    general — seulement celles-ci, et la conformite du resultat est
    **revalidee contre le motif d'origine** avant d'etre rendue.
    """
    entete = re.match(r"^\^(\d+)\((.+)\)\$$", motif)
    if not entete:
        return []
    indicatif, alternatives = entete.group(1), entete.group(2)

    variantes: list[tuple[str, list[_Segment]]] = []
    for alternative in alternatives.split("|"):
        segments: list[_Segment] = []
        position = 0
        while position < len(alternative):
            reste = alternative[position:]
            if classe := re.match(r"^\[([\d\-]+)\]", reste):
                segments.append((_chiffres_de_la_classe(classe.group(1)), 1))
                position += classe.end()
            elif repetition := re.match(r"^\\d\{(\d+)\}", reste):
                segments.append((None, int(repetition.group(1))))
                position += repetition.end()
            elif litteral := re.match(r"^\d+", reste):
                segments.append(litteral.group(0))
                position += litteral.end()
            else:  # pragma: no cover — motif hors grammaire connue
                return []
        variantes.append((indicatif, segments))
    return variantes


@dataclass(frozen=True, slots=True)
class Devise:
    """Devise, avec la zone monetaire qui la porte.

    Le point decisif : `pays` vient de la colonne `countries_using`. La devise
    n'est donc **pas** un choix libre parmi deux — elle est **determinee par le
    pays**. XAF est la zone CEMAC (Cameroun), XOF la zone UEMOA (Cote
    d'Ivoire, Burkina Faso, Senegal). Emettre XOF pour un client camerounais
    serait aussi faux qu'emettre une devise inventee, et aucun service ne le
    refuserait (FRA-222).
    """

    code: str
    nom: str
    decimales: int
    banque_centrale: str
    pays: frozenset[str]


@dataclass(slots=True)
class RapportGeographique:
    """Rapport de couverture, exige par EF-06 en debut d'execution."""

    pays: list[str] = field(default_factory=list)
    nb_regions: int = 0
    nb_villes: int = 0
    nb_quartiers: int = 0
    nb_telcos: int = 0
    nb_devises: int = 0
    orphelins: list[str] = field(default_factory=list)

    def resume(self) -> str:
        lignes = [
            f"Pays      : {len(self.pays)} ({', '.join(sorted(self.pays))})",
            f"Regions   : {self.nb_regions}",
            f"Villes    : {self.nb_villes}",
            f"Quartiers : {self.nb_quartiers}",
            f"Telcos    : {self.nb_telcos}",
            f"Devises   : {self.nb_devises}",
            "Chaine    : pays -> devise -> telco -> region -> ville -> quartier",
        ]
        if self.orphelins:
            lignes.append(f"Exclus pour rattachement invalide : {len(self.orphelins)}")
            lignes.extend(f"  - {motif}" for motif in self.orphelins)
        return "\n".join(lignes)


class ReferentielGeo:
    """Arbre geographique a 5 niveaux, indexe pour un acces direct (UC-05)."""

    def __init__(
        self,
        regions: dict[str, Region],
        villes: dict[str, City],
        quartiers: dict[str, District],
        rapport: RapportGeographique,
        telcos: dict[str, Telco] | None = None,
        devises: dict[str, Devise] | None = None,
        pays_index: dict[str, Pays] | None = None,
    ) -> None:
        self.regions = regions
        self.villes = villes
        self.quartiers = quartiers
        self.rapport = rapport
        self.telcos = telcos or {}
        self.devises = devises or {}
        self.pays_index = pays_index or {}

    # -- Telecoms et monnaie — la coherence que le systeme ne verifie pas ----

    def telcos_du_pays(self, pays: str) -> list[Telco]:
        """Operateurs d'un pays, du plus gros au plus petit."""
        code = str(pays).strip().upper()
        return sorted(
            (t for t in self.telcos.values() if t.country_iso2 == code),
            key=lambda t: t.part_marche,
            reverse=True,
        )

    def operateur_du_msisdn(self, msisdn: str, pays: str) -> Telco | None:
        """`EF-27` — a quel operateur du pays ce numero appartient-il ?

        Renvoie `None` si aucun ne le reconnait : le numero n'est alors
        attribuable a aucun reseau reel du pays. Aucun service FinZuu ne fait
        cette verification.
        """
        for telco in self.telcos_du_pays(pays):
            if telco.accepte(msisdn):
                return telco
        return None

    def tirer_telco(self, pays: str, alea: random.Random) -> Telco | None:
        """Tire un operateur **selon sa part de marche reelle**.

        MTN CM 46 %, Orange CM 43 %, Camtel 3 % — au Senegal, Orange pese
        55 %. Repartir les 2000 clients uniformement entre trois operateurs
        produirait un echantillon qui ne ressemble a aucun marche africain, et
        un bailleur qui connait la zone le verrait au premier graphique.

        Le referentiel porte ces parts depuis le depart ; nous ne les
        utilisions pas.
        """
        operateurs = self.telcos_du_pays(pays)
        if not operateurs:
            return None
        total = sum(t.part_marche for t in operateurs)
        if total <= 0:
            return alea.choice(operateurs)
        tirage = alea.uniform(0, total)
        cumul = 0.0
        for operateur in operateurs:
            cumul += operateur.part_marche
            if tirage <= cumul:
                return operateur
        return operateurs[-1]

    def composer_msisdn(self, pays: str, chiffres: str, alea: random.Random) -> tuple[str, Telco]:
        """Compose un MSISDN conforme au plan reel du pays, et rend l'operateur.

        **La doctrine du Loader, appliquee a un cinquieme champ.** Faker donne
        huit chiffres sans le prefixe operateur, qui est implicite pour un
        habitant du pays — au Cameroun tout mobile commence par `6`, en Cote
        d'Ivoire par `01`, `05` ou `07`, au Burkina par `0`. Sans lui, aucun
        numero de Faker n'est attribuable a un reseau reel (mesure du 09/08).

        La matiere reste celle de Faker ; c'est le prefixe, et lui seul, que
        nous ajoutons. Exactement comme pour les raisons sociales, les dates de
        naissance, les adresses et les occupations.
        """
        operateur = self.tirer_telco(pays, alea)
        if operateur is None:
            raise ValueError(f"aucun operateur telecom pour le pays {pays!r} — EF-27 inapplicable")
        return operateur.composer_msisdn(chiffres, alea), operateur

    def pays(self, code: str) -> Pays | None:
        """Fiche pays complete — indicatif, devise, TVA, fuseau, regulateurs."""
        return self.pays_index.get(str(code).strip().upper())

    def indicatif(self, code: str) -> str:
        """L'indicatif telephonique, depuis le referentiel et non code en dur."""
        fiche = self.pays(code)
        if fiche is None or not fiche.dial_code:
            raise ValueError(f"aucun indicatif pour le pays {code!r}")
        return fiche.dial_code

    def tva_du_pays(self, code: str) -> float:
        """Taux de TVA — 19,25 % au Cameroun, 18 % en zone UEMOA.

        La Policy d'un produit Collecte porte un champ `vat`. Il doit venir
        d'ici, pas d'une constante inventee.
        """
        fiche = self.pays(code)
        if fiche is None:
            raise ValueError(f"pays inconnu : {code!r}")
        return fiche.tva_percent

    def devise_du_pays(self, code: str) -> Devise | None:
        """La devise n'est pas un choix : elle est **declaree par le pays**.

        Deux chemins existent dans le referentiel — `Countries.currency_iso` et
        `Currencies.countries_using`. On suit le premier, qui est direct, et le
        chargement verifie que les deux concordent. Une divergence est
        journalisee comme orpheline plutot que silencieusement arbitree.
        """
        fiche = self.pays(code)
        if fiche is None:
            return None
        return self.devises.get(fiche.devise_iso)

    # -- Acces indexes ------------------------------------------------------

    def region(self, region_id: str) -> Region | None:
        """Acces direct par identifiant — renvoie None plutot que de lever.

        Un identifiant absent du referentiel ne doit jamais interrompre une
        generation : l'appelant retombe sur l'identifiant brut et le rapport
        garde la trace. Meme principe que `parse_datetime`.
        """
        return self.regions.get(region_id)

    def ville(self, city_id: str) -> City | None:
        return self.villes.get(city_id)

    def quartier(self, district_id: str) -> District | None:
        return self.quartiers.get(district_id)

    def regions_du_pays(self, pays: str) -> list[Region]:
        return sorted(
            (r for r in self.regions.values() if r.country_iso2 == pays),
            key=lambda r: r.region_id,
        )

    def villes_de_region(self, region_id: str) -> list[City]:
        return sorted(
            (v for v in self.villes.values() if v.region_id == region_id),
            key=lambda v: v.city_id,
        )

    def quartiers_de_ville(self, city_id: str) -> list[District]:
        return sorted(
            (q for q in self.quartiers.values() if q.city_id == city_id),
            key=lambda q: q.district_id,
        )

    def villes_porteuses_de_quartiers(self, pays: str) -> list[City]:
        """Seules villes ou un Kiosque peut exister.

        UC-09, exception : « Si le referentiel geographique ne contient aucun
        district pour une ville, aucun Kiosque n'est cree dans cette ville. »
        C'est donc cette liste, et non la liste complete des villes, qui borne
        le nombre d'Agences utiles d'un pays.
        """
        avec_quartiers = {q.city_id for q in self.quartiers.values()}
        return sorted(
            (
                v
                for v in self.villes.values()
                if v.country_iso2 == pays and v.city_id in avec_quartiers
            ),
            key=lambda v: (-v.poids_economique, v.city_id),
        )

    def nb_quartiers_du_pays(self, pays: str) -> int:
        villes_pays = {v.city_id for v in self.villes.values() if v.country_iso2 == pays}
        return sum(1 for q in self.quartiers.values() if q.city_id in villes_pays)


def charger_referentiel(chemin: Path) -> ReferentielGeo:
    """Lit le classeur et valide l'emboitement (EF-01, EF-02).

    Exception EF-01 / UC-05 : si le fichier est introuvable, l'execution est
    interrompue avec le chemin attendu explicitement journalise — jamais un
    echec silencieux.
    """
    if not chemin.exists():
        raise FileNotFoundError(
            f"Referentiel geographique introuvable. Chemin attendu : {chemin.resolve()}"
        )

    classeur = openpyxl.load_workbook(chemin, read_only=True, data_only=True)
    rapport = RapportGeographique()

    # Pays — la racine. Douze colonnes, dont dial_code, currency_iso et
    # tva_percent, que nous ignorions entierement.
    pays_index: dict[str, Pays] = {}
    for ligne in _lignes(classeur, "Countries"):
        code = str(ligne.get("country_iso2") or "").strip().upper()
        if not code:
            continue
        pays_index[code] = Pays(
            iso2=code,
            nom_fr=str(ligne.get("country_name_fr") or ""),
            nom_en=str(ligne.get("country_name_en") or ""),
            capitale=str(ligne.get("capital") or ""),
            dial_code=str(ligne.get("dial_code") or "").strip(),
            devise_iso=str(ligne.get("currency_iso") or "").strip().upper(),
            tva_percent=_flottant(ligne.get("tva_percent")),
            timezone=str(ligne.get("timezone") or ""),
            region_africa=str(ligne.get("region_africa") or ""),
            regulateur_telco=str(ligne.get("regulator_telco") or ""),
            regulateur_finance=str(ligne.get("regulator_finance") or ""),
        )
    pays_connus = set(pays_index)
    rapport.pays = sorted(pays_connus)

    regions: dict[str, Region] = {}
    for ligne in _lignes(classeur, "Regions"):
        pays = str(ligne.get("country_iso2") or "").strip().upper()
        region_id = str(ligne.get("region_id") or "").strip()
        if not region_id:
            continue
        if pays not in pays_connus:
            rapport.orphelins.append(f"Region {region_id} : pays {pays!r} inconnu")
            continue
        regions[region_id] = Region(
            region_id=region_id,
            country_iso2=pays,
            name=str(ligne.get("region_name") or ""),
            capitale=str(ligne.get("region_capital") or ""),
            population=_entier_ou_none(ligne.get("population_est")),
        )

    villes: dict[str, City] = {}
    for ligne in _lignes(classeur, "Cities"):
        city_id = str(ligne.get("city_id") or "").strip()
        region_id = str(ligne.get("region_id") or "").strip()
        if not city_id:
            continue
        if region_id not in regions:
            rapport.orphelins.append(f"Ville {city_id} : region {region_id!r} inconnue")
            continue
        villes[city_id] = City(
            city_id=city_id,
            country_iso2=regions[region_id].country_iso2,
            region_id=region_id,
            name=str(ligne.get("city_name") or ""),
            poids_economique=_flottant(ligne.get("weight_economic")),
            latitude=_flottant_ou_none(ligne.get("latitude")),
            longitude=_flottant_ou_none(ligne.get("longitude")),
            population=_entier_ou_none(ligne.get("population_est")),
            est_capitale_pays=bool(ligne.get("is_capital_country")),
            est_capitale_region=bool(ligne.get("is_capital_region")),
        )

    quartiers: dict[str, District] = {}
    for ligne in _lignes(classeur, "Districts"):
        district_id = str(ligne.get("district_id") or "").strip()
        city_id = str(ligne.get("city_id") or "").strip()
        if not district_id:
            continue
        if city_id not in villes:
            rapport.orphelins.append(f"Quartier {district_id} : ville {city_id!r} inconnue")
            continue
        quartiers[district_id] = District(
            district_id=district_id,
            city_id=city_id,
            name=str(ligne.get("district_name") or ""),
            zone_type=str(ligne.get("zone_type") or "").strip().lower(),
            population=_entier_ou_none(ligne.get("population_est")),
        )

    # Telcos — 12 operateurs avec leur regex de numerotation (EF-27) et leur
    # part de marche reelle. Ces donnees etaient dans le referentiel depuis le
    # depart ; nous ne les chargions pas.
    telcos: dict[str, Telco] = {}
    for ligne in _lignes(classeur, "Telcos"):
        telco_id = str(ligne.get("telco_id") or "").strip()
        pays = str(ligne.get("country_iso2") or "").strip().upper()
        motif = str(ligne.get("regex_msisdn") or "").strip()
        if not telco_id:
            continue
        if pays not in pays_connus:
            rapport.orphelins.append(f"Telco {telco_id} : pays {pays!r} inconnu")
            continue
        if not motif:
            rapport.orphelins.append(f"Telco {telco_id} : aucun regex_msisdn — EF-27 impossible")
            continue
        try:
            re.compile(motif)
        except re.error as erreur:
            rapport.orphelins.append(f"Telco {telco_id} : regex invalide ({erreur})")
            continue
        telcos[telco_id] = Telco(
            telco_id=telco_id,
            country_iso2=pays,
            network_name=str(ligne.get("network_name") or ""),
            short_name=str(ligne.get("short_name") or ""),
            ussd_base_code=str(ligne.get("ussd_base_code") or ""),
            regex_msisdn=motif,
            part_marche=_flottant(ligne.get("market_share_pct")),
        )

    # Devises — la zone monetaire determine la devise du pays, elle ne se
    # choisit pas. `countries_using` porte l'appartenance.
    devises: dict[str, Devise] = {}
    for ligne in _lignes(classeur, "Currencies"):
        code = str(ligne.get("currency_iso") or "").strip().upper()
        if not code:
            continue
        utilisateurs = {
            morceau.strip().upper()
            for morceau in str(ligne.get("countries_using") or "").split(",")
            if morceau.strip()
        }
        devises[code] = Devise(
            code=code,
            nom=str(ligne.get("currency_name") or ""),
            decimales=_entier_ou_none(ligne.get("decimal_places")) or 0,
            banque_centrale=str(ligne.get("central_bank") or ""),
            pays=frozenset(utilisateurs),
        )

    classeur.close()

    # EF-02 etendu : un pays sans operateur ou sans devise rend impossible la
    # generation d'un client credible. On le signale plutot que de le decouvrir
    # au 500e onboarding.
    for code in sorted(pays_connus):
        fiche = pays_index[code]
        if not any(t.country_iso2 == code for t in telcos.values()):
            rapport.orphelins.append(f"Pays {code} : aucun operateur telecom — EF-27 inapplicable")
        if not fiche.dial_code:
            rapport.orphelins.append(f"Pays {code} : aucun indicatif telephonique")
        if fiche.devise_iso not in devises:
            rapport.orphelins.append(
                f"Pays {code} : devise declaree {fiche.devise_iso!r} absente de la feuille "
                "Currencies — la chaine pays -> devise est rompue"
            )
        elif code not in devises[fiche.devise_iso].pays:
            rapport.orphelins.append(
                f"Pays {code} : declare {fiche.devise_iso}, mais cette devise ne le liste pas "
                "dans countries_using — les deux feuilles se contredisent"
            )
        if not any(r.country_iso2 == code for r in regions.values()):
            rapport.orphelins.append(f"Pays {code} : aucune region")

    rapport.nb_regions = len(regions)
    rapport.nb_villes = len(villes)
    rapport.nb_quartiers = len(quartiers)
    rapport.nb_telcos = len(telcos)
    rapport.nb_devises = len(devises)
    return ReferentielGeo(regions, villes, quartiers, rapport, telcos, devises, pays_index)


def _lignes(classeur: Any, feuille: str) -> list[dict[str, Any]]:
    """Lit une feuille en dictionnaires indexes par en-tete."""
    if feuille not in classeur.sheetnames:
        return []
    iterateur = classeur[feuille].iter_rows(values_only=True)
    try:
        entetes = [str(cellule).strip() if cellule else "" for cellule in next(iterateur)]
    except StopIteration:
        return []
    resultat: list[dict[str, Any]] = []
    for ligne in iterateur:
        if all(cellule is None for cellule in ligne):
            continue
        resultat.append(dict(zip(entetes, ligne, strict=False)))
    return resultat


def _flottant(valeur: Any) -> float:
    try:
        return float(valeur)
    except (TypeError, ValueError):
        return 0.0


def _flottant_ou_none(valeur: Any) -> float | None:
    """EF-03 : « lorsqu'elles sont disponibles ». Une coordonnee absente reste
    None — jamais 0.0, qui designerait le golfe de Guinee."""
    try:
        return float(valeur)
    except (TypeError, ValueError):
        return None


def _entier_ou_none(valeur: Any) -> int | None:
    try:
        return int(float(valeur))
    except (TypeError, ValueError):
        return None
